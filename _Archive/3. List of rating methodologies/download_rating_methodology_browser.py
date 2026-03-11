import argparse
import re
import shutil
import tempfile
import time
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

from openpyxl import load_workbook
import win32com.client as win32

# ---------------- Configuration ----------------

EXCEL_PATH = Path(
    r"C:\Users\michael.lewellyn\OneDrive - Bancorp\Desktop\Moody's Score Card\3. List of rating methodologies\List of Rating Methodologies (21-10-2025).xlsx"
)
DOWNLOAD_DIR = Path(
    r"C:\Users\michael.lewellyn\OneDrive - Bancorp\Desktop\Moody's Score Card\3. List of rating methodologies\Moody's Methodologies"
)
SHEET_NAME = "Rating Methodologies"
START_ROW = 12
MAX_ROW = 179

# Use Chrome by default (it’s installed on your machine and tends to behave
# more predictably with Selenium in many environments). If you prefer Edge,
# set BROWSER="edge" or pass --browser edge on the command line.
BROWSER = "chrome"  # "edge" | "chrome"

# IMPORTANT: this must match your actual logged-in browser profile.
# Edge default:
EDGE_USER_DATA_DIR = Path(r"C:\Users\michael.lewellyn\AppData\Local\Microsoft\Edge\User Data")
EDGE_PROFILE = "Default"

# Chrome default:
CHROME_USER_DATA_DIR = Path(r"C:\Users\michael.lewellyn\AppData\Local\Google\Chrome\User Data")
CHROME_PROFILE = "Default"

# How long to wait for a download to complete after clicking the download button.
DOWNLOAD_TIMEOUT_S = 90


# ---------------- Helpers ----------------


def _safe_filename(candidate: str, fallback: str) -> str:
    """Return a filesystem-safe filename that ends in .pdf."""
    base = (candidate or "").strip() or fallback
    base = re.sub(r'[<>:"/\\|?*]', "_", base)
    if not base.lower().endswith(".pdf"):
        base = f"{base}.pdf"
    return base


def _infer_filename(cell_value: str, url: str) -> str:
    tail = Path(urlparse(url).path).name
    fallback = tail or "rating_methodology"
    return _safe_filename(cell_value or fallback, fallback)


def _read_excel_links(start_row: int, max_row: int):
    """
    Yield (row, display_name, url) from the workbook.

    We prefer reading hyperlinks directly instead of driving Excel via COM to avoid
    OLE errors and to keep this script fully headless.
    """
    source_path = EXCEL_PATH
    try:
        # Need full cells (including hyperlinks). read_only=True returns ReadOnlyCell
        # objects that don't expose hyperlink targets.
        wb = load_workbook(source_path, read_only=False, data_only=True)
    except PermissionError:
        # Common on OneDrive / when the workbook is open in Excel: file is locked.
        # Best effort: copy to a temp file and read from there.
        tmp = Path(tempfile.mkdtemp(prefix="moodys_xlsx_")) / source_path.name
        try:
            shutil.copy2(source_path, tmp)
        except Exception as exc:
            raise PermissionError(
                f"Permission denied reading Excel workbook:\n  {source_path}\n\n"
                "Fix:\n"
                "- Close the workbook in Excel (and ensure Excel isn't holding a lock), OR\n"
                "- Copy the .xlsx out of OneDrive to a local folder and rerun.\n\n"
                f"Copy attempt to temp also failed: {exc}"
            ) from exc

        print(f"Excel workbook appears locked; reading from temp copy: {tmp}")
        wb = load_workbook(tmp, read_only=False, data_only=True)
    ws = wb[SHEET_NAME]

    for row in range(start_row, max_row + 1):
        cell = ws[f"A{row}"]
        value = str(cell.value).strip() if cell.value is not None else ""

        hyperlink = cell.hyperlink.target if cell.hyperlink else None
        if not hyperlink:
            # Stop when we hit a blank row and no link.
            if not value:
                break
            continue

        yield row, value, hyperlink


def _choose_browser_profile() -> tuple[str, Path, str]:
    if BROWSER.lower() == "chrome":
        return "chrome", CHROME_USER_DATA_DIR, CHROME_PROFILE
    return "edge", EDGE_USER_DATA_DIR, EDGE_PROFILE


def _looks_like_pdf(path: Path) -> bool:
    try:
        return path.read_bytes()[:4] == b"%PDF"
    except Exception:
        return False


def _handle_save_as_dialog(full_destination: Path, timeout_seconds: int = 15) -> bool:
    """
    If the browser shows a Windows "Save As" dialog, fill the file name + click Save.
    Returns True if we handled a dialog, else False.
    """
    try:
        from pywinauto import Desktop
        from pywinauto.keyboard import send_keys
    except Exception:  # pragma: no cover
        return False

    end = time.time() + timeout_seconds
    while time.time() < end:
        try:
            # There can be multiple "Save As" windows; pick the active one if possible.
            dlgs = Desktop(backend="uia").windows(title_re=".*Save As.*", visible_only=True)
            if not dlgs:
                time.sleep(0.2)
                continue
            dlg = None
            for d in dlgs:
                try:
                    if d.has_focus() or d.is_active():
                        dlg = d
                        break
                except Exception:
                    continue
            if dlg is None:
                dlg = dlgs[0]

            dlg.set_focus()

            # Prefer setting the filename via UIA edit control (more reliable than hotkeys),
            # but keep hotkey fallback for different dialog layouts.
            wrote_path = False
            try:
                edit = dlg.child_window(control_type="Edit")
                if edit.exists(timeout=0.5):
                    edit.set_focus()
                    send_keys("^a")
                    send_keys(str(full_destination), with_spaces=True)
                    wrote_path = True
            except Exception:
                wrote_path = False

            if not wrote_path:
                send_keys("%n")  # focus File name
                send_keys("^a")
                send_keys(str(full_destination), with_spaces=True)

            # Click Save (or press Enter).
            clicked_save = False
            try:
                # Button text can vary slightly ("Save", "&Save"), so keep it loose.
                save_btn = dlg.child_window(title_re=".*Save.*", control_type="Button")
                if save_btn.exists(timeout=0.8):
                    save_btn.click_input()
                    clicked_save = True
            except Exception:
                clicked_save = False

            if not clicked_save:
                # Enter usually activates the default action (Save).
                send_keys("{ENTER}")
                time.sleep(0.2)
                # Alt+S is also a common accelerator for Save on Windows dialogs.
                if _save_as_dialog_exists():
                    try:
                        send_keys("%s")
                    except Exception:
                        pass

            # Overwrite confirmation may appear.
            try:
                confirm = Desktop(backend="uia").window(title_re=".*Confirm Save As.*")
                if confirm.exists(timeout=1):
                    confirm.set_focus()
                    try:
                        yes_btn = confirm.child_window(
                            title_re="^(Yes|Replace)$", control_type="Button"
                        )
                        if yes_btn.exists(timeout=0.5):
                            yes_btn.click_input()
                        else:
                            send_keys("%y")
                    except Exception:
                        send_keys("%y")
            except Exception:
                pass

            return True
        except Exception:
            time.sleep(0.2)
            continue

    return False


def _save_as_dialog_exists() -> bool:
    try:
        from pywinauto import Desktop

        dlg = Desktop(backend="uia").window(title_re=".*Save As.*")
        return bool(dlg.exists(timeout=0.1))
    except Exception:
        return False


def _wait_for_new_pdf(download_dir: Path, before: set[Path], timeout_seconds: int) -> Path:
    """
    Wait for a new .pdf file to appear in download_dir (and not be a partial .crdownload).
    """
    end = time.time() + timeout_seconds
    while time.time() < end:
        partials = list(download_dir.glob("*.crdownload"))
        if partials:
            time.sleep(0.25)
            continue

        after = set(download_dir.glob("*.pdf"))
        created = sorted(after - before, key=lambda p: p.stat().st_mtime, reverse=True)
        if created:
            return created[0]
        time.sleep(0.25)

    raise TimeoutError(
        f"No new PDF appeared in {download_dir} within {timeout_seconds}s after clicking download."
    )


def _follow_excel_hyperlink_via_com(worksheet, row: int) -> tuple[str, str] | None:
    """
    Mimic your manual behavior: click/follow the hyperlink from Excel itself.

    Returns (display_name, url) if there is a hyperlink in A{row}, else None.
    """
    cell = worksheet.Range(f"A{row}")
    value = str(cell.Value).strip() if cell.Value is not None else ""
    if cell.Hyperlinks.Count == 0:
        if not value:
            return None
        return None

    link = cell.Hyperlinks(1)
    url = link.Address
    if not url:
        return None

    worksheet.Activate()
    cell.Select()
    # This is the programmatic equivalent of clicking the hyperlink.
    link.Follow(NewWindow=False, AddHistory=True)
    return value, url


def _save_current_browser_tab(destination: Path, timeout_seconds: int) -> None:
    """
    Save the currently focused browser tab via Ctrl+S and handle the Save As dialog.

    This avoids DOM automation entirely and works best when you are already logged in
    (same session as when you click the link manually from Excel).
    """
    from pywinauto.keyboard import send_keys

    # Trigger Save for the active tab (for PDF viewer this saves the PDF).
    send_keys("^s")

    handled = _handle_save_as_dialog(destination, timeout_seconds=20)
    if not handled:
        raise RuntimeError(
            "Save As dialog did not appear after Ctrl+S. "
            "If Chrome is configured to auto-download without prompting, "
            "use the Selenium mode instead."
        )

    end = time.time() + timeout_seconds
    while time.time() < end:
        if destination.exists() and _looks_like_pdf(destination):
            return
        time.sleep(0.25)
    raise TimeoutError(f"Timed out waiting for saved PDF to appear: {destination}")


def _get_default_downloads_dir() -> Path:
    """
    Return the actual Downloads folder path on Windows (handles OneDrive redirection).

    We use the Shell namespace rather than assuming ~/Downloads.
    """
    try:
        import win32com.client as _win32  # local import to keep this helper self-contained

        shell = _win32.Dispatch("Shell.Application")
        downloads = shell.NameSpace("shell:Downloads")
        if downloads is not None:
            p = Path(downloads.Self.Path)
            if p.exists():
                return p
    except Exception:
        pass

    # Fallback: common default.
    p = Path.home() / "Downloads"
    return p if p.exists() else Path.home()


def _wait_for_new_pdf_in_dir(download_dir: Path, before: set[Path], timeout_seconds: int) -> Path:
    end = time.time() + timeout_seconds
    while time.time() < end:
        partials = list(download_dir.glob("*.crdownload"))
        if partials:
            time.sleep(0.25)
            continue

        after = set(download_dir.glob("*.pdf"))
        created = sorted(after - before, key=lambda p: p.stat().st_mtime, reverse=True)
        if created:
            return created[0]
        time.sleep(0.25)
    raise TimeoutError(f"No new PDF appeared in {download_dir} within {timeout_seconds}s.")


def _wait_for_new_pdf_in_either_dir(
    dir_a: Path,
    before_a: set[Path],
    dir_b: Path,
    before_b: set[Path],
    timeout_seconds: int,
) -> Path:
    """
    Wait for a new PDF to appear in either directory. Returns the path to the new file.
    """
    end = time.time() + timeout_seconds
    while time.time() < end:
        # Ignore in-progress Chrome partials in either location.
        if list(dir_a.glob("*.crdownload")) or list(dir_b.glob("*.crdownload")):
            time.sleep(0.25)
            continue

        after_a = set(dir_a.glob("*.pdf"))
        created_a = sorted(after_a - before_a, key=lambda p: p.stat().st_mtime, reverse=True)
        if created_a:
            return created_a[0]

        after_b = set(dir_b.glob("*.pdf"))
        created_b = sorted(after_b - before_b, key=lambda p: p.stat().st_mtime, reverse=True)
        if created_b:
            return created_b[0]

        time.sleep(0.25)

    raise TimeoutError(
        f"No new PDF appeared in either {dir_a} or {dir_b} within {timeout_seconds}s."
    )


def _click_viewer_download_button(timeout_seconds: int = 15) -> None:
    """
    Click the 'Download' button in the currently active browser window using UI Automation.

    This is intentionally DOM-agnostic so it works even when the PDF toolbar is inside an iframe.
    """
    from pywinauto import Desktop

    end = time.time() + timeout_seconds
    last_err: Optional[Exception] = None
    # Be specific to avoid matching Excel/other Moody windows.
    title_re = r".*(ratings\.moodys\.com|Ratings\.Moodys\.com).*"

    while time.time() < end:
        try:
            desk = Desktop(backend="uia")
            wins = desk.windows(title_re=title_re, visible_only=True)
            if not wins:
                time.sleep(0.5)
                continue

            # Choose the most likely active browser window.
            win = None
            for w in wins:
                try:
                    if w.has_focus() or w.is_active():
                        win = w
                        break
                except Exception:
                    continue
            if win is None:
                win = wins[0]

            win.set_focus()

            # Search for any button-like control labeled Download.
            for ctrl_type in ("Button", "MenuItem", "Hyperlink"):
                try:
                    for el in win.descendants(control_type=ctrl_type):
                        text = (el.window_text() or "").strip()
                        if text and re.search(r"download", text, re.I):
                            el.click_input()
                            time.sleep(0.6)
                            if _save_as_dialog_exists():
                                return
                            return
                except Exception as exc:
                    last_err = exc

            # Fallback: click near the top-right toolbar area where the download icon usually sits.
            try:
                rect = win.rectangle()
                candidates = [
                    (rect.right - 140, rect.top + 145),
                    (rect.right - 110, rect.top + 145),
                    (rect.right - 80, rect.top + 145),
                    (rect.right - 140, rect.top + 160),
                    (rect.right - 110, rect.top + 160),
                    (rect.right - 80, rect.top + 160),
                ]
                for x, y in candidates:
                    win.click_input(coords=(x - rect.left, y - rect.top))
                    time.sleep(0.6)
                    if _save_as_dialog_exists():
                        return
                return
            except Exception as exc:
                last_err = exc

            time.sleep(0.5)
        except Exception as exc:
            last_err = exc
            time.sleep(0.5)

    raise RuntimeError(f"Could not click the viewer Download button. Last error: {last_err}")


def _auto_download_from_viewer(destination: Path, timeout_seconds: int) -> None:
    """
    Excel-mode helper:

    - We open the methodology page via Excel hyperlink (uses your already logged-in session).
    - We click the download icon in the viewer (like you do manually).
    - Then we either handle the Save As dialog (preferred) or detect a file landing in Downloads.

    This avoids the "login prompt" caused by Selenium temp profiles and avoids Ctrl+S saving
    the HTML container page ("Webpage, Complete").
    """
    downloads_dir = _get_default_downloads_dir()
    print(f"Downloads folder detected as: {downloads_dir}")
    before_downloads = set(downloads_dir.glob("*.pdf"))
    dest_dir = destination.parent
    dest_dir.mkdir(parents=True, exist_ok=True)
    before_dest = set(dest_dir.glob("*.pdf"))

    _click_viewer_download_button(timeout_seconds=15)

    # If Save As appears, force our destination; otherwise accept auto-download to Downloads.
    print("Clicked Download; waiting for Save As dialog...")
    handled = _handle_save_as_dialog(destination, timeout_seconds=30)
    if handled:
        end = time.time() + timeout_seconds
        while time.time() < end:
            if destination.exists() and _looks_like_pdf(destination):
                return
            time.sleep(0.25)
        raise TimeoutError(f"Save As handled but {destination} did not appear as a valid PDF.")

    # Auto-download: wait for the new PDF in Downloads and move/rename to destination.
    print("No Save As dialog detected; waiting for PDF to appear in Downloads...")
    downloaded = _wait_for_new_pdf_in_either_dir(
        downloads_dir,
        before_downloads,
        dest_dir,
        before_dest,
        timeout_seconds=timeout_seconds,
    )

    # If it landed directly in the destination directory, just validate and return.
    if downloaded.parent.resolve() == dest_dir.resolve():
        if downloaded != destination:
            try:
                if destination.exists():
                    destination.unlink()
            except Exception:
                pass
            downloaded.replace(destination)
        if not _looks_like_pdf(destination):
            raise RuntimeError(f"Downloaded file is not a valid PDF: {destination}")
        return

    try:
        if destination.exists():
            destination.unlink()
    except Exception:
        pass
    downloaded.replace(destination)

    end = time.time() + timeout_seconds
    while time.time() < end:
        if destination.exists() and _looks_like_pdf(destination):
            return
        time.sleep(0.25)
    raise TimeoutError(f"Timed out waiting for saved PDF to appear: {destination}")


def _create_webdriver(
    download_dir: Path,
    browser: str,
    user_data_dir: Path,
    profile: str,
    driver_path: Optional[Path] = None,
    use_temp_profile: bool = False,
):
    """
    Start Chrome/Edge via Selenium with a deterministic download directory and (optionally)
    your existing logged-in profile.

    Goal: click the viewer download icon and have the PDF land in DOWNLOAD_DIR without
    manual intervention.
    """
    download_dir.mkdir(parents=True, exist_ok=True)

    prefs = {
        "download.default_directory": str(download_dir),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        # Keep PDFs in the browser viewer so the site toolbar exists:
        "plugins.always_open_pdf_externally": False,
    }

    # NOTE: We avoid webdriver-manager by default because many corporate networks block
    # its driver download hosts (e.g. msedgedriver.azureedge.net). Selenium's built-in
    # "Selenium Manager" can often locate a local driver without downloading; if it can't,
    # we support passing a driver executable explicitly via --driver-path.
    from selenium import webdriver

    browser_l = browser.lower()
    temp_profile_dir: Optional[Path] = None
    if use_temp_profile:
        temp_profile_dir = Path(tempfile.mkdtemp(prefix="moodys_selenium_profile_"))

    if browser_l == "chrome":
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        from selenium.webdriver.chrome.service import Service as ChromeService

        options = ChromeOptions()
        options.add_argument("--disable-gpu")
        options.add_argument("--start-maximized")
        options.add_experimental_option("prefs", prefs)

        if use_temp_profile and temp_profile_dir is not None:
            options.add_argument(f"--user-data-dir={str(temp_profile_dir)}")
        else:
            options.add_argument(f"--user-data-dir={str(user_data_dir)}")
            options.add_argument(f"--profile-directory={profile}")

        if driver_path is not None:
            return webdriver.Chrome(options=options, service=ChromeService(str(driver_path)))

        # Let Selenium Manager resolve the driver (may still download on some machines).
        return webdriver.Chrome(options=options)

    # Edge (Chromium)
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.edge.service import Service as EdgeService

    options = EdgeOptions()
    options.use_chromium = True
    options.add_argument("--disable-gpu")
    options.add_argument("--start-maximized")
    options.add_experimental_option("prefs", prefs)

    if use_temp_profile and temp_profile_dir is not None:
        options.add_argument(f"--user-data-dir={str(temp_profile_dir)}")
    else:
        options.add_argument(f"--user-data-dir={str(user_data_dir)}")
        options.add_argument(f"--profile-directory={profile}")

    if driver_path is not None:
        return webdriver.Edge(options=options, service=EdgeService(str(driver_path)))

    # Let Selenium Manager resolve the driver (may still download on some machines).
    return webdriver.Edge(options=options)


def _click_download_anywhere(driver) -> None:
    """
    Click the PDF viewer download icon.

    Moody's uses a PDF viewer toolbar; we try a few robust selectors.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    candidates = [
        (By.CSS_SELECTOR, "button[aria-label='Download']"),
        (By.CSS_SELECTOR, "a[aria-label='Download']"),
        (By.CSS_SELECTOR, "button[title='Download']"),
        (By.CSS_SELECTOR, "a[title='Download']"),
        (By.CSS_SELECTOR, "[aria-label*='download' i]"),
        (By.CSS_SELECTOR, "[title*='download' i]"),
        (By.CSS_SELECTOR, "button#download"),
        (By.CSS_SELECTOR, "a#download"),
        (By.CSS_SELECTOR, ".toolbarButton.download"),
        (By.CSS_SELECTOR, "[data-testid*='download']"),
    ]

    last_err: Optional[Exception] = None

    def try_click_in_current_context() -> bool:
        nonlocal last_err
        for by, sel in candidates:
            try:
                btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((by, sel)))
                driver.execute_script("arguments[0].click();", btn)
                return True
            except Exception as exc:
                last_err = exc
        return False

    # Try main document first.
    if try_click_in_current_context():
        return

    # Then try any iframes (viewer toolbars are sometimes embedded).
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
    except Exception:
        iframes = []

    for iframe in iframes:
        try:
            driver.switch_to.frame(iframe)
            if try_click_in_current_context():
                driver.switch_to.default_content()
            return
        except Exception as exc:
            last_err = exc
        finally:
            try:
                driver.switch_to.default_content()
            except Exception:
                pass

    raise RuntimeError(f"Could not find/click the download button. Last error: {last_err}")


# ---------------- Main ----------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Download Moody's rating methodologies by opening each link in a real browser "
            "and clicking the PDF viewer download button."
        )
    )
    parser.add_argument("--start-row", type=int, default=START_ROW)
    parser.add_argument("--end-row", type=int, default=MAX_ROW)
    parser.add_argument(
        "--mode",
        choices=["excel", "selenium"],
        default="excel",
        help=(
            "excel: click/follow hyperlinks from Excel (uses your existing logged-in browser session), "
            "then save via Ctrl+S + Save As. "
            "selenium: open URLs directly in an automated browser and click Download."
        ),
    )
    parser.add_argument("--browser", choices=["edge", "chrome"], default=BROWSER)
    parser.add_argument(
        "--driver-path",
        type=str,
        default="",
        help=(
            "Optional: path to msedgedriver.exe or chromedriver.exe. "
            "Use this if your network blocks Selenium/driver downloads."
        ),
    )
    parser.add_argument(
        "--temp-profile",
        action="store_true",
        help=(
            "Use a temporary fresh browser profile (avoids Edge/Chrome profile lock crashes). "
            "If you use this, you may need to log in to ratings.moodys.com again."
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional: stop after downloading N files (0 means no limit).",
    )
    parser.add_argument(
        "--page-load-wait",
        type=float,
        default=5.0,
        help="Seconds to wait after opening the PDF page before saving/downloading.",
    )
    args = parser.parse_args()

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    if args.mode == "excel":
        excel = None
        workbook = None
        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = True
            workbook = excel.Workbooks.Open(str(EXCEL_PATH), UpdateLinks=False, ReadOnly=True)
            worksheet = workbook.Worksheets(SHEET_NAME)
            worksheet.Activate()

            downloaded = 0
            for row in range(args.start_row, args.end_row + 1):
                result = _follow_excel_hyperlink_via_com(worksheet, row)
                if result is None:
                    continue
                name, url = result

                filename = _infer_filename(name, url)
                destination = DOWNLOAD_DIR / filename
                if destination.exists() and destination.stat().st_size > 0:
                    print(f"Row {row}: already exists, skipping: {destination.name}")
                    continue

                print(f"Row {row}: opened via Excel link; waiting {args.page_load_wait}s for page to load...")
                time.sleep(float(args.page_load_wait))

                # IMPORTANT: don't use Ctrl+S here — it often saves the HTML page ("Webpage, Complete").
                # Instead, click the viewer's Download icon and handle Save As / Downloads automatically.
                _auto_download_from_viewer(destination, timeout_seconds=DOWNLOAD_TIMEOUT_S)
                print(f"Row {row}: saved {destination.name}")
                downloaded += 1
                if args.limit and downloaded >= args.limit:
                    break

        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
        return

    # Override browser choice from CLI, then re-resolve profile paths.
    browser_name, user_data_dir, profile = (
        ("chrome", CHROME_USER_DATA_DIR, CHROME_PROFILE)
        if args.browser.lower() == "chrome"
        else ("edge", EDGE_USER_DATA_DIR, EDGE_PROFILE)
    )
    if not user_data_dir.exists():
        raise RuntimeError(
            f"Browser user data dir not found: {user_data_dir}\n"
            f"Update EDGE_USER_DATA_DIR/CHROME_USER_DATA_DIR to match your machine."
        )

    driver_path = Path(args.driver_path).expanduser() if args.driver_path else None
    try:
        driver = _create_webdriver(
            DOWNLOAD_DIR,
            browser=browser_name,
            user_data_dir=user_data_dir,
            profile=profile,
            driver_path=driver_path,
            use_temp_profile=args.temp_profile,
        )
    except Exception as exc:
        # Common failure mode on corporate machines: Edge/Chrome crashes at launch due to
        # profile locks or managed policies. If user didn't request temp profile, retry
        # once with a temporary profile automatically.
        msg = str(exc)
        should_retry_temp = (not args.temp_profile) and (
            "DevToolsActivePort" in msg or "session not created" in msg.lower()
        )
        if should_retry_temp:
            print(
                "Browser failed to start with your normal profile; retrying once with a temporary profile..."
            )
            driver = _create_webdriver(
                DOWNLOAD_DIR,
                browser=browser_name,
                user_data_dir=user_data_dir,
                profile=profile,
                driver_path=driver_path,
                use_temp_profile=True,
            )
        else:
            raise RuntimeError(
                "Failed to start the browser for automation.\n"
                "- Close ALL Edge/Chrome windows first (including background/tray), then retry.\n"
                "- Or run with --temp-profile to avoid profile lock issues.\n"
                "- If driver downloads are blocked, pass --driver-path to a local msedgedriver.exe/chromedriver.exe.\n"
                f"Original error: {exc}"
            ) from exc
    try:
        driver.set_page_load_timeout(120)
        downloaded = 0
        for row, name, url in _read_excel_links(args.start_row, args.end_row):
            filename = _infer_filename(name, url)
            destination = DOWNLOAD_DIR / filename

            if destination.exists() and destination.stat().st_size > 0:
                print(f"Row {row}: already exists, skipping: {destination.name}")
                continue

            downloads_dir = _get_default_downloads_dir()
            before_downloads = set(downloads_dir.glob("*.pdf"))
            before_dest = set(DOWNLOAD_DIR.glob("*.pdf"))
            print(f"Row {row}: opening {url}")
            driver.get(url)
            time.sleep(float(args.page_load_wait))  # allow viewer toolbar to settle (Moody's pages can be slow)

            _click_download_anywhere(driver)

            # If Save As appears, handle it; otherwise wait for PDF in Downloads OR destination.
            handled = _handle_save_as_dialog(destination)
            if handled:
                end = time.time() + DOWNLOAD_TIMEOUT_S
                while time.time() < end:
                    if destination.exists() and _looks_like_pdf(destination):
                        break
                    time.sleep(0.25)
                else:
                    raise TimeoutError(f"Save As handled but {destination} did not appear as a valid PDF.")
            else:
                downloaded_path = _wait_for_new_pdf_in_either_dir(
                    downloads_dir,
                    before_downloads,
                    DOWNLOAD_DIR,
                    before_dest,
                    timeout_seconds=DOWNLOAD_TIMEOUT_S,
                )
                if downloaded_path != destination:
                    try:
                        if destination.exists():
                            destination.unlink()
                    except Exception:
                        pass
                    downloaded_path.replace(destination)

                if not _looks_like_pdf(destination):
                    raise RuntimeError(f"Downloaded file is not a valid PDF: {destination}")

            print(f"Row {row}: saved {destination.name}")
            downloaded += 1
            if args.limit and downloaded >= args.limit:
                break
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()


